# ==============================================================================
# SISTEMA MESTRE: CÁLCULO ARC FLASH + RELATÓRIO + ADESIVO (INPUTS VALIDADOS)
# ==============================================================================
# Descrição:
# 1. Realiza cálculos IEEE 1584 (Blindado contra erros de digitação e zeros).
# 2. Coleta dados Administrativos (CEP/CNPJ validados).
# 3. Gera Relatório em Word (.docx).
# 4. Gera Adesivo em Excel (.xlsx) e converte para PDF (A1:J15 Centralizado).
# ==============================================================================

import math
import os
import sys
import datetime
import locale
import subprocess
from datetime import datetime as dt

# --- 1. PREPARAÇÃO (Instalação Rápida) ---
def install_and_import(package):
    try:
        __import__(package)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])

install_and_import("docxtpl")
install_and_import("openpyxl")

from docxtpl import DocxTemplate, RichText
import openpyxl
from openpyxl.cell.cell import MergedCell
from google.colab import drive

# Instalação do LibreOffice (Silenciosa) para PDF
chk = subprocess.run("which soffice", shell=True, stdout=subprocess.PIPE)
if chk.returncode != 0:
    print(">> Instalando recursos de PDF (aguarde)...")
    os.system("apt-get update -qq && apt-get install -y libreoffice-core libreoffice-calc --no-install-recommends -qq > /dev/null")

# Configuração de Data
try: locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except: pass

# --- 2. CONFIGURAÇÃO ---
PASTA_DRIVE = '/content/drive/MyDrive/Colab Notebooks/EEI + RL + AD/'
MODELO_DOC = "ESTUDO DE ENERGIA INCIDENTE - MODELO.docx"
MODELO_XLS = "ADESIVO ENERGIA INCIDENTE - MODELO.xlsx"

if not os.path.exists('/content/drive'):
    drive.mount('/content/drive')

# ==============================================================================
# PARTE A: TABELAS E CONSTANTES (IEEE 1584-2018)
# ==============================================================================
TABLE_1 = {'VCB': [-0.04287, 1.035, -0.083, 0, 0, -4.783e-9, 1.962e-6, -0.000229, 0.003141, 1.092], 'VCBB': [-0.017432, 0.98, -0.05, 0, 0, -5.767e-9, 2.524e-6, -0.00034, 0.01187, 1.013], 'HCB': [0.054922, 0.988, -0.11, 0, 0, -5.382e-9, 2.316e-6, -0.000302, 0.0091, 0.9725], 'VOA': [0.043785, 1.04, -0.18, 0, 0, -4.783e-9, 1.962e-6, -0.000229, 0.003141, 1.092], 'HOA': [0.111147, 1.008, -0.24, 0, 0, -3.895e-9, 1.641e-6, -0.000197, 0.002615, 1.1]}
TABLE_2 = {'VCB': [0, -1.4269e-6, 8.3137e-5, -0.0019382, 0.022366, -0.12645, 0.30226], 'VCBB': [1.138e-6, -6.0287e-5, 0.0012758, -0.013778, 0.080217, -0.24066, 0.33524], 'HCB': [0, -3.097e-6, 0.00016405, -0.0033609, 0.033308, -0.16182, 0.34627], 'VOA': [9.5606e-7, -5.1543e-5, 0.0011161, -0.01242, 0.075125, -0.23584, 0.33696], 'HOA': [0, -3.1555e-6, 0.0001682, -0.0034607, 0.034124, -0.1599, 0.34629]}
TABLE_3 = {'VCB': [0.753364, 0.566, 1.752636, 0, 0, -4.783e-9, 1.962e-6, -0.000229, 0.003141, 1.092, 0, -1.598, 0.957], 'VCBB': [3.068459, 0.26, -0.098107, 0, 0, -5.767e-9, 2.524e-6, -0.00034, 0.01187, 1.013, -0.06, -1.809, 1.19], 'HCB': [4.073745, 0.344, -0.370259, 0, 0, -5.382e-9, 2.316e-6, -0.000302, 0.0091, 0.9725, 0, -2.03, 1.036], 'VOA': [0.679294, 0.746, 1.222636, 0, 0, -4.783e-9, 1.962e-6, -0.000229, 0.003141, 1.092, 0, -1.598, 0.997], 'HOA': [3.470417, 0.465, -0.261863, 0, 0, -3.895e-9, 1.641e-6, -0.000197, 0.002615, 1.1, 0, -1.99, 1.04]}
TABLE_7_TYPICAL = {'VCB': [-0.000302, 0.03441, 0.4325], 'VCBB': [-0.0002976, 0.032, 0.479], 'HCB': [-0.0001923, 0.01935, 0.6899]}
TABLE_7_SHALLOW = {'VCB': [0.002222, -0.02556, 0.6222], 'VCBB': [-0.002778, 0.1194, -0.2778], 'HCB': [-0.0005556, 0.03722, 0.4778]}
CONSTANTS_AB = {'VCB':  {'A': 4,  'B': 20}, 'VCBB': {'A': 10, 'B': 24}, 'HCB':  {'A': 10, 'B': 22}}

# ==============================================================================
# PARTE B: FUNÇÕES AUXILIARES
# ==============================================================================
def log10(x):
    if x <= 0: return 0
    return math.log10(x)

# Input numérico float (Protege contra letras e Zeros onde proibido)
def get_float_input(prompt, permitir_zero=False):
    while True:
        try:
            valor = float(input(prompt).replace(',', '.').strip())
            if not permitir_zero and valor <= 0:
                print("   ❌ O valor deve ser maior que zero.")
                continue
            return valor
        except ValueError:
            print("   ❌ Entrada inválida. Digite um número.")

# [NOVO] Input para CEP e CNPJ (Valida dígitos e tamanho)
def get_input_numerico(prompt, tamanho_exato):
    while True:
        val = input(prompt).strip()
        # Limpa pontuação se o usuário digitar (ex: 38.360-000 vira 38360000)
        val_limpo = val.replace('.', '').replace('-', '').replace('/', '')

        if not val_limpo.isdigit():
            print("   ❌ Digite apenas números.")
            continue

        if len(val_limpo) != tamanho_exato:
            print(f"   ❌ Deve conter exatamente {tamanho_exato} dígitos (Digitado: {len(val_limpo)}).")
            continue

        return val_limpo

def print_separator(title=None, char="="):
    print("\n\n" + char*75)
    if title: print(f" {title.upper()}")
    print(char*75)

def print_result(label, value, unit="", decimals=3):
    if isinstance(value, str): val_str = value
    elif decimals == 0: val_str = f"{value:.0f}"
    else: val_str = f"{value:.{decimals}f}"
    print(f" > {label:<35} : {val_str} {unit}")

def print_energy(label, e_cal, e_joule):
    val_str = f"{e_cal:.3f} cal/cm² ({e_joule:.3f} J/cm²)"
    print(f" > {label:<35} : {val_str}")

def obter_categoria_nfpa(e): return ("Isento (< 1.2 cal)", "Não requer AR") if e <= 1.2 else ("Categoria 1", "4.0 cal/cm²") if e <= 4.0 else ("Categoria 2", "8.0 cal/cm²") if e <= 8.0 else ("Categoria 3", "25.0 cal/cm²") if e <= 25.0 else ("Categoria 4", "40.0 cal/cm²")
def obter_zonas_choque_nr10(v): return (0, 0) if v < 50 else (200, 700) if v <= 1000 else (700, 1500)
def calcular_ajuste_linear(d, v, c): A, B = CONSTANTS_AB[c]['A'], CONSTANTS_AB[c]['B']; return (660.4 + (d - 660.4) * ((v + A) / B)) * (1 / 25.4) if c in CONSTANTS_AB else d / 25.4
def calcular_ees_correto(C, H, W, D, V):
    H_in, W_in = H/25.4, W/25.4
    if V < 0.6 and H < 508 and W < 508 and D <= 203.2: return (H_in+W_in)/2, "Shallow (Raso)", H_in, W_in
    W1 = 20.0 if W < 508 else W_in if W <= 660.4 else calcular_ajuste_linear(W if W <= 1244.6 else 1244.6, V, C)
    H1 = (20.0 if H < 508 else H_in if H <= 1244.6 else 49.0) if C == 'VCB' else (20.0 if H < 508 else H_in if H <= 660.4 else calcular_ajuste_linear(H if H <= 1244.6 else 1244.6, V, C))
    return (H1 + W1)/2, "Typical (Típico)", H1, W1

# --- AUXILIARES FORMATAÇÃO ---
def fmt_br(v, c=2): return f"{v:.{c}f}".replace('.', ',')
def formatar_cep(v): s="".join(filter(str.isdigit,str(v))); return f"{s[:2]}.{s[2:5]}-{s[5:]}" if len(s)==8 else v
def formatar_cnpj(v): s="".join(filter(str.isdigit,str(v))); return f"{s[:2]}.{s[2:5]}.{s[5:8]}/{s[8:12]}-{s[12:]}" if len(s)==14 else v

def formatar_coeficientes_rt(dic):
    rt = RichText(); items = list(dic.items())
    for i, (k, v) in enumerate(items):
        l = "".join([c for c in k if not c.isdigit()]); n = "".join([c for c in k if c.isdigit()])
        rt.add(l, font='Times New Roman', size=24, italic=True)
        rt.add(n, font='Times New Roman', size=24, italic=True, subscript=True)
        val = f"{v:.4f}".replace('.', ',') if abs(v) >= 0.0001 or v == 0 else f"{v:.2e}".replace('.', ',')
        rt.add(f" = {val}", font='Times New Roman', size=24, italic=True)
        if i < len(items)-1: rt.add("; ", font='Times New Roman', size=24, italic=True)
    return rt

def converter_para_pdf(caminho_arquivo, pasta_saida="/content/"):
    cmd = f"soffice --headless --convert-to pdf --outdir '{pasta_saida}' '{caminho_arquivo}' > /dev/null 2>&1"
    os.system(cmd)

# --- EXCEL ---
def writing_excel(ws, r, c, v):
    cell = ws.cell(row=r, column=c)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if cell.coordinate in mr: ws[mr.start_cell.coordinate].value = v; return
    else: cell.value = v

def fill_excel_label(ws, lbl, v, off):
    for r in ws.iter_rows(min_row=1, max_row=25, min_col=1, max_col=12):
        for c in r:
            if c.value and isinstance(c.value, str) and lbl.lower() in str(c.value).lower():
                writing_excel(ws, c.row, c.column+off, v); return

# ==============================================================================
# FLUXO PRINCIPAL
# ==============================================================================
def executar_calculo():
    print("CALCULADORA DE ENERGIA INCIDENTE (IEEE 1584-2018) - 208 <= V <= 600V\n")

    nome_eq = input(">> Nome do Equipamento: ").strip()

    Voc_V = get_float_input(">> Tensão (Voc) [V]: ", permitir_zero=False)
    Voc = Voc_V / 1000.0
    if Voc > 0.6: print("(!) AVISO: Cálculo somente para 208<= Voc <= 600V.")

    Ibf = get_float_input(">> Corrente de Curto (Ibf) [kA]: ", permitir_zero=False)

    while True:
        Config = input(">> Configuração Eletrodos (VCB, VCBB, HCB, VOA, HOA): ").strip().upper()
        if Config in TABLE_1: break
        print(f"   ❌ Configuração '{Config}' inválida. Tente novamente.")

    is_open_air = Config in ['VOA', 'HOA']
    if is_open_air: H, W, D = 0.0, 0.0, 0.0
    else:
        H = get_float_input(">> Altura Painel [mm]: ", permitir_zero=True)
        W = get_float_input(">> Largura Painel [mm]: ", permitir_zero=True)
        D = get_float_input(">> Profundidade Painel [mm]: ", permitir_zero=True)

    Gap = get_float_input(">> Distância Entre Condutores (G) [mm]: ", permitir_zero=False)
    Dist = get_float_input(">> Distância de Trabalho (D) [mm]: ", permitir_zero=False)

    # --- CÁLCULOS ---
    k = TABLE_1[Config]
    term1 = 10 ** (k[0] + k[1]*log10(Ibf) + k[2]*log10(Gap))
    term2 = (k[3]*Ibf**6 + k[4]*Ibf**5 + k[5]*Ibf**4 + k[6]*Ibf**3 + k[7]*Ibf**2 + k[8]*Ibf + k[9])
    Iarc600 = term1 * term2

    term_a = (0.6/Voc)**2
    term_b = (1/Iarc600)**2 - ((0.6**2 - Voc**2)/(0.6**2 * Ibf**2))

    if term_a * term_b <= 0:
        print("\n❌ ERRO MATEMÁTICO: Raiz negativa detectada (combinação Voc/Ibf inválida).")
        return None

    Iarc = 1 / math.sqrt(term_a * term_b)

    print_separator("CÁLCULOS INTERMEDIÁRIOS", "-")
    print_result("Corrente de Arco Elétrico em 600V (Iarc_600)", Iarc600, "kA")
    print_result("Corrente Final de Arco Elétrico (Iarc)", Iarc, "kA")

    T_ms = get_float_input(">> Tempo de atuação para Iarc (T) [ms]: ", permitir_zero=False)
    print("")

    if is_open_air: CF, box_type, EES, b = 1.0, "Open Air (Ar Livre)", 0.0, [0,0,0]
    else:
        EES, box_type, H1, W1 = calcular_ees_correto(Config, H, W, D, Voc)
        b = TABLE_7_SHALLOW[Config] if "Shallow" in box_type else TABLE_7_TYPICAL[Config]
        CF = 1/(b[0]*EES**2 + b[1]*EES + b[2]) if "Shallow" in box_type else b[0]*EES**2 + b[1]*EES + b[2]

    vk = TABLE_2[Config]
    VarCf = vk[0]*Voc**6 + vk[1]*Voc**5 + vk[2]*Voc**4 + vk[3]*Voc**3 + vk[4]*Voc**2 + vk[5]*Voc + vk[6]
    Imin = Iarc * (1 - 0.5 * VarCf)

    print(f" > {'Tipo de Invólucro':<35} : {box_type}")
    if not is_open_air:
        print_result("Dimensão Equivalente (H1 x W1)", f"{H1:.2f} x {W1:.2f}", "in")
        print_result("Tamanho Equivalente (EES)", EES, "in")
    print_result("Fator de Correção (CF)", CF, "\n")
    print_result("Corrente Reduzida (Iarc_min)", Imin, "kA")

    T_min = get_float_input(">> Tempo de duração para Iarc_min (T_min) [ms]: ", permitir_zero=False)

    tk = TABLE_3[Config]
    C2 = tk[3]*Ibf**7 + tk[4]*Ibf**6 + tk[5]*Ibf**5 + tk[6]*Ibf**4 + tk[7]*Ibf**3 + tk[8]*Ibf**2 + tk[9]*Ibf
    C3 = tk[10]*log10(Ibf) + tk[11]*log10(Dist) + log10(1/CF)

    E_cal = ((12.552/50)*T_ms*(10**(tk[0] + tk[1]*log10(Gap) + (tk[2]*Iarc600/C2) + C3 + tk[12]*log10(Iarc))))/4.184
    AFB = Dist * (1.2/E_cal)**(1/tk[11])
    E_min_cal = ((12.552/50)*T_min*(10**(tk[0] + tk[1]*log10(Gap) + (tk[2]*Iarc600/C2) + C3 + tk[12]*log10(Imin))))/4.184
    AFB_min = Dist * (1.2/E_min_cal)**(1/tk[11])

    # Convertendo para Joules para exibição
    E_joule = E_cal * 4.184
    E_min_joule = E_min_cal * 4.184

    print_separator("CENÁRIO 1: CORRENTE NOMINAL", "-")
    print_energy("Energia Incidente", E_cal, E_joule)
    print_result("Distância Segura de Aproximação - AFB", AFB, "mm", 0)

    print_separator("CENÁRIO 2: CORRENTE REDUZIDA", "-")
    print_energy("Energia Incidente (Reduzida)", E_min_cal, E_min_joule)
    print_result("Distância Segura de Aprox. - AFB (Reduzida)", AFB_min, "mm", 0)

    E_final = max(E_cal, E_min_cal)
    AFB_final = max(AFB, AFB_min)
    cat_epi, rating = obter_categoria_nfpa(E_final)
    zr, zc = obter_zonas_choque_nr10(Voc_V)

    # Textos do pior caso
    pior_caso_txt = "Corrente Nominal" if E_final == E_cal else "Corrente Reduzida"
    E_final_joule = max(E_joule, E_min_joule)

    print_separator("RESULTADO FINAL DO ESTUDO", "=")
    print_result("Cenário Pior Caso", pior_caso_txt, "")
    print_energy("Energia Incidente", E_final, E_final_joule)
    print_result("Distância Segura de Aprox. - AFB", AFB_final, "mm", 0)
    print("-" * 75)
    print_result("Categoria EPI (NFPA 70E)", cat_epi, "")
    print_result("Suportabilidade Mínima", rating, "")
    print("="*75 + "\n")

    return {
        'equipamento': nome_eq, 'energia': E_final, 'dist_trabalho': Dist, 'afb': AFB_final,
        'cat_epi': cat_epi, 'categoria_risco': cat_epi, 'rating': rating, 'atpv': rating,
        'tensao': Voc_V, 'zc': zc, 'zr': zr, 'config': Config, 'ibf': Ibf, 'gap': Gap, 'dist': Dist,
        'dimensoes': f"{H:.0f}x{W:.0f}x{D:.0f} mm" if not is_open_air else "Ar Livre",
        'tipo_involucro': box_type, 'ia_600': Iarc600, 'i_arc': Iarc, 'tempo': T_ms, 'ees': EES, 'cf': CF,
        'e_nominal': E_cal, 'afb_nominal': AFB, 'var_cf': VarCf, 'i_min': Imin, 'tempo_min': T_min,
        'e_min': E_min_cal, 'afb_min': AFB_min, 'e_final': E_final, 'afb_final': AFB_final,
        'dict_step2': {f'k{i+1}':v for i,v in enumerate(k)}, 'dict_step5': {f'b{i+1}':v for i,v in enumerate(b)} if not is_open_air else {},
        'dict_step6': {f'k{i+1}':v for i,v in enumerate(tk)}, 'dict_step8': {f'k{i+1}':v for i,v in enumerate(vk)}
    }

def gerar_arquivos(d):
    print("\n--- DADOS DO RELATÓRIO ---")
    cli = input("Nome do Cliente: ") or "CLIENTE PADRÃO"
    end = input("Endereço: ") or "Rua da Indústria, 100"

    # [VALIDAÇÃO CEP E CNPJ]
    cep_input = get_input_numerico("CEP (apenas números): ", 8)
    cnpj_input = get_input_numerico("CNPJ (apenas números): ", 14)

    print("\n>> Gerando relatório...")
    mes = {1:'Janeiro', 2:'Fevereiro', 3:'Março', 4:'Abril', 5:'Maio', 6:'Junho', 7:'Julho', 8:'Agosto', 9:'Setembro', 10:'Outubro', 11:'Novembro', 12:'Dezembro'}[dt.now().month]

    # 1. WORD
    ctx = {
        'mês_hoje': mes, 'data_hoje': dt.now().strftime("%d/%m/%Y"),
        'nome_equipamento': d['equipamento'], 'nome_cliente': cli, 'endereço_cliente': end,
        'cep_cliente': formatar_cep(cep_input), 'cnpj_cliente': formatar_cnpj(cnpj_input),
        'config': d['config'], 'voc': fmt_br(d['tensao'],0), 'ibf': fmt_br(d['ibf']), 'gap': fmt_br(d['gap'],1), 'dist': fmt_br(d['dist'],0),
        'dimensoes': d['dimensoes'], 'tipo_involucro': d['tipo_involucro'], 'ia_600': fmt_br(d['ia_600'],3),
        'i_arc': fmt_br(d['i_arc'],3), 'tempo': fmt_br(d['tempo'],1), 'ees': fmt_br(d['ees']), 'cf': fmt_br(d['cf'],3),
        'e_nominal': fmt_br(d['e_nominal']), 'afb_nominal': fmt_br(d['afb_nominal'],0), 'var_cf': fmt_br(d['var_cf'],3),
        'i_min': fmt_br(d['i_min'],3), 'tempo_min': fmt_br(d['tempo_min'],1), 'e_min': fmt_br(d['e_min']), 'afb_min': fmt_br(d['afb_min'],0),
        'e_final': fmt_br(d['e_final']), 'afb_final': fmt_br(d['afb_final'],0), 'categoria_risco': d['categoria_risco'], 'atpv': d['atpv'],
        'coeficientes_step2': formatar_coeficientes_rt(d.get('dict_step2')),
        'txt_coeficientes_step5': formatar_coeficientes_rt(d.get('dict_step5')),
        'coeficientes_step5': formatar_coeficientes_rt(d.get('dict_step5')),
        'coeficientes_step6': formatar_coeficientes_rt(d.get('dict_step6')),
        'coeficientes_step8': formatar_coeficientes_rt(d.get('dict_step8')), 'k_dist': " "
    }
    try:
        doc = DocxTemplate(os.path.join(PASTA_DRIVE, MODELO_DOC))
        doc.render(ctx)
        doc.save(f"EEI - {d['equipamento']}.docx")
        print("✅ SUCESSO! Relatório DOCX gerado.")
    except Exception as e: print(f"❌ Erro Word: {e}")

    # 2. EXCEL
    print(">> Gerando adesivo...")
    try:
        wb = openpyxl.load_workbook(os.path.join(PASTA_DRIVE, MODELO_XLS))
        ws = wb.active

        fill_excel_label(ws, "Energia incidente", f"{fmt_br(d['energia'])} cal/cm²", 2)
        fill_excel_label(ws, "Limite do arco", f"{fmt_br(d['afb'],0)} mm", 2)
        fill_excel_label(ws, "Distância de trabalho", f"{fmt_br(d['dist_trabalho'],0)} mm", 2)
        fill_excel_label(ws, "Categoria de risco", d['cat_epi'], 2)
        fill_excel_label(ws, "Suportabilidade mínima", str(d['rating']).replace('.',','), 2)
        cls = "00 (≤ 500 V)" if d['tensao'] <= 500 else "0 (≤ 1000 V)"
        fill_excel_label(ws, "Classe:", cls, 1)
        fill_excel_label(ws, "Tensão:", f"{fmt_br(d['tensao'],0)} V", 1)
        fill_excel_label(ws, "Zona controlada:", f"{fmt_br(d['zc'],0)} mm", 1)
        fill_excel_label(ws, "Zona de risco:", f"{fmt_br(d['zr'],0)} mm", 1)
        fill_excel_label(ws, "Equipamento:", f"Equipamento: {d['equipamento']}", 0)
        try: writing_excel(ws, 13, 9, dt.now().strftime('%b/%Y').upper())
        except: pass

        try:
            ws.print_area = 'A1:J15'
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToHeight = 1
            ws.page_setup.fitToWidth = 1
            ws.page_setup.horizontalCentered = True
            ws.page_setup.verticalCentered = True
        except: pass

        xls_name = f"ADESIVO EEI - {d['equipamento'].replace(' ','_')}.xlsx"
        wb.save(xls_name)
        print(f"✅ SUCESSO! Adesivo XLSX gerado.")

        print(">> Imprimindo adesivo (PDF)...")
        converter_para_pdf(xls_name)
        print(f"✅ SUCESSO! PDF gerado: {xls_name.replace('.xlsx', '.pdf')}")

    except Exception as e: print(f"❌ Erro Excel: {e}")

if __name__ == "__main__":
    res = executar_calculo()
    if res:
        gerar_arquivos(res)
        print("\n" + "="*50)
        print(" PROCESSO FINALIZADO! OS ARQUIVOS JÁ PODEM SER BAIXADOS.")
        print("="*75)
